"""Where election data lives on disk, and the Excel/CSV writers.

Everything lands under <Desktop>/C-LABS Digital EVM/, never inside the app's
own folder - that keeps the packaged .exe read-only and means OneDrive (on
Windows) or Time Machine (on macOS) can back up the one folder that matters.
"""

import csv
import ctypes
import os
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

APP_FOLDER_NAME = "C-LABS Digital EVM"


def desktop_path():
    """Resolve the real Desktop folder, accounting for Windows OneDrive redirection.

    `Path.home() / "Desktop"` is wrong on a Windows machine where OneDrive has
    redirected Desktop elsewhere - votes would silently land in a stale,
    unsynced folder. SHGetKnownFolderPath(FOLDERID_Desktop) asks Windows for
    the *actual* current path. macOS/Linux have no such redirection, so
    `~/Desktop` is reliable there.
    """
    if sys.platform == "win32":
        FOLDERID_Desktop = "{B4BFCC3A-DB2C-424C-B029-7FE99A87C641}"
        buf = ctypes.c_wchar_p()
        guid = ctypes.create_unicode_buffer(FOLDERID_Desktop)
        # SHGetKnownFolderPath wants a GUID struct, not a string; build it via ctypes.wintypes.
        import ctypes.wintypes as wintypes

        class GUID(ctypes.Structure):
            _fields_ = [
                ("Data1", ctypes.c_ulong), ("Data2", ctypes.c_ushort),
                ("Data3", ctypes.c_ushort), ("Data4", ctypes.c_ubyte * 8),
            ]

        rfid = GUID(0xB4BFCC3A, 0xDB2C, 0x424C,
                    (ctypes.c_ubyte * 8)(0xB0, 0x29, 0x7F, 0xE9, 0x9A, 0x87, 0xC6, 0x41))
        path_ptr = ctypes.c_wchar_p()
        result = ctypes.windll.shell32.SHGetKnownFolderPath(
            ctypes.byref(rfid), 0, 0, ctypes.byref(path_ptr)
        )
        if result == 0 and path_ptr.value:
            return Path(path_ptr.value)
        return Path.home() / "Desktop"  # fallback, should not happen on real Windows
    return Path.home() / "Desktop"


def app_root():
    root = desktop_path() / APP_FOLDER_NAME
    (root / "Database").mkdir(parents=True, exist_ok=True)
    (root / "Excel").mkdir(parents=True, exist_ok=True)
    (root / "Backup").mkdir(parents=True, exist_ok=True)
    return root


def db_path():
    return app_root() / "Database" / "election.db"


def journal_path():
    return app_root() / "Backup" / "journal.csv"


def votes_xlsx_path():
    return app_root() / "Excel" / "votes.xlsx"


def append_journal(serial, post_title, candidate_name, is_test):
    """Crash-safe, human-readable append-only log - one line per (ballot, post)."""
    path = journal_path()
    is_new = not path.exists()
    with open(path, "a", newline="") as f:
        writer = csv.writer(f)
        if is_new:
            writer.writerow(["Time", "Vote ID", "Position", "Candidate", "Test"])
        writer.writerow([
            time.strftime("%Y-%m-%d %H:%M:%S"),
            "%03d" % serial,
            post_title,
            candidate_name or "NOTA",
            "TEST" if is_test else "",
        ])
        f.flush()
        os.fsync(f.fileno())


def write_results_xlsx(tally_rows, ballot_count, path=None):
    """Regenerate votes.xlsx from the current tally - cheap enough to do on every cast."""
    path = path or votes_xlsx_path()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Post", "Candidate", "Votes"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for post in tally_rows:
        for cand in post["candidates"]:
            ws.append([post["post_title"], cand["name"], cand["count"]])
        ws.append([post["post_title"], "NOTA", post["nota"]])

    summary = wb.create_sheet("Summary")
    summary.append(["Total ballots cast", ballot_count])
    summary["A1"].font = Font(bold=True)

    tmp_path = str(path) + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, path)  # atomic on both platforms - never a half-written file
    return path


def backup_snapshot():
    """Copy election.db and votes.xlsx into Backup/ with a timestamped name."""
    import shutil

    root = app_root()
    stamp = time.strftime("%Y%m%d-%H%M%S")
    backup_dir = root / "Backup"
    src_db = root / "Database" / "election.db"
    src_xlsx = root / "Excel" / "votes.xlsx"
    if src_db.exists():
        shutil.copy2(src_db, backup_dir / ("backup-%s.db" % stamp))
    if src_xlsx.exists():
        shutil.copy2(src_xlsx, backup_dir / ("backup-%s.xlsx" % stamp))
    return backup_dir
